"""
Script to add ALL baskets from mock_data to Excel
This will ensure hide/unhide filter works properly
"""
import sys
sys.path.append('data')
from mock_data import baskets_data, funds_data
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

EXCEL_PATH = 'data/AlphaniftyMasterData.xlsx'

def update_excel_with_all_baskets():
    """Add all baskets from mock_data to Excel"""
    
    print(f"📊 Updating Excel with ALL {len(baskets_data)} baskets...")
    
    # Prepare data from mock_data
    baskets_list = []
    fund_allocations_list = []
    sector_allocations_list = []
    top_holdings_list = []
    
    for basket in baskets_data:
        # Add basket
        baskets_list.append({
            'BasketID': basket['id'],
            'BasketName': basket['name'],
            'Color': basket['color'],
            'Description': basket['description'],
            'AgeRange': basket['ageRange'],
            'RiskLevel': basket['riskLevel'],
            'MinInvestment': basket['minInvestment'],
            'TimeHorizon': basket['timeHorizon'],
            'Goals': ', '.join(basket['goals']),
            'ExperienceLevel': basket['experienceLevel'],
            'CAGR_1Y': basket.get('cagr1Y', 0),
            'CAGR_3Y': basket.get('cagr3Y', 0),
            'CAGR_5Y': basket.get('cagr5Y', 0),
            'RiskPercentage': basket.get('riskPercentage', 0),
            'SharpeRatio': basket.get('sharpeRatio', 0),
            'ExpectedReturn': basket.get('expectedReturn', 0),
            'Rationale': basket.get('rationale', ''),
            'Philosophy': basket.get('philosophy', ''),
            'SuitableFor': basket.get('suitableFor', ''),
            'RebalancingFrequency': basket.get('rebalancingFrequency', 'Quarterly'),
            'ExcelFile': basket.get('excelFile', ''),
            'InfoFile': basket.get('infoFile', '')
        })
        
        # Add fund allocations
        if 'fundAllocations' in basket and basket['fundAllocations']:
            for fund_alloc in basket['fundAllocations']:
                fund_allocations_list.append({
                    'BasketID': basket['id'],
                    'FundID': fund_alloc['fundId'],
                    'FundName': fund_alloc['fundName'],
                    'AllocationPercent': fund_alloc['allocationPercent'],
                    'Category': fund_alloc.get('category', ''),
                    'Returns1Y': fund_alloc.get('returns1Y', 0),
                    'Returns3Y': fund_alloc.get('returns3Y', 0),
                    'Returns5Y': fund_alloc.get('returns5Y', 0),
                    'Corpus': fund_alloc.get('corpus', ''),
                    'NAV': fund_alloc.get('nav', 0)
                })
        
        # Add sector allocations
        if 'sectorAllocation' in basket and basket['sectorAllocation']:
            for sector in basket['sectorAllocation']:
                sector_allocations_list.append({
                    'BasketID': basket['id'],
                    'Sector': sector['sector'],
                    'AllocationPercent': sector['percent']
                })
        
        # Add top holdings
        if 'topHoldings' in basket and basket['topHoldings']:
            for holding in basket['topHoldings']:
                top_holdings_list.append({
                    'BasketID': basket['id'],
                    'StockName': holding.get('stockName', ''),
                    'AllocationPercent': holding.get('percent', 0),
                    'Sector': holding.get('sector', ''),
                    'Type': holding.get('type', 'Equity')
                })
    
    # Create DataFrames
    df_baskets = pd.DataFrame(baskets_list)
    df_fund_alloc = pd.DataFrame(fund_allocations_list)
    df_sector_alloc = pd.DataFrame(sector_allocations_list)
    df_top_holdings = pd.DataFrame(top_holdings_list)
    
    # Get existing Funds sheet (we don't change this)
    df_funds_existing = pd.read_excel(EXCEL_PATH, sheet_name='Funds')
    
    print(f"✅ Prepared data:")
    print(f"   - Baskets: {len(df_baskets)} rows")
    print(f"   - Fund Allocations: {len(df_fund_alloc)} rows")
    print(f"   - Sector Allocations: {len(df_sector_alloc)} rows")
    print(f"   - Top Holdings: {len(df_top_holdings)} rows")
    print(f"   - Funds: {len(df_funds_existing)} rows (existing)")
    
    # Write to Excel
    print("\n📝 Writing to Excel...")
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='w') as writer:
        df_baskets.to_excel(writer, sheet_name='Baskets', index=False)
        df_fund_alloc.to_excel(writer, sheet_name='FundAllocations', index=False)
        df_sector_alloc.to_excel(writer, sheet_name='SectorAllocations', index=False)
        df_top_holdings.to_excel(writer, sheet_name='TopHoldings', index=False)
        df_funds_existing.to_excel(writer, sheet_name='Funds', index=False)
    
    # Format Excel
    print("🎨 Formatting Excel...")
    wb = load_workbook(EXCEL_PATH)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(EXCEL_PATH)
    
    print(f"\n✅ SUCCESS! Excel updated with ALL {len(df_baskets)} baskets!")
    print(f"📄 File: {EXCEL_PATH}")
    print(f"\n🎯 Now the hide/unhide filter will show all baskets!")
    
    # Show basket IDs
    basket_ids = sorted([b['id'] for b in baskets_data])
    print(f"\n📋 All baskets in Excel:")
    print(f"   {', '.join(basket_ids)}")

if __name__ == '__main__':
    update_excel_with_all_baskets()
