"""
Complete Excel Template Generator for Alphanifty - 8 Valid Baskets
Excludes: Yellow (b4), White (b11), Dusshera (b15), and baskets without Excel files
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_complete_excel():
    """Create complete Excel with all 8 valid baskets and their data"""
    
    excel_path = os.path.join(os.path.dirname(__file__), 'AlphaniftyMasterData_NEW.xlsx')
    
    # Sheet 1: Baskets - All 8 valid baskets
    baskets_data = {
        'BasketID': ['b9', 'b10', 'b12', 'b13', 'b14', 'b16', 'b17', 'b18'],
        'Name': [
            'Aggressive Premium',
            'Conservative Balanced Basket',
            'Every Common India',
            'Raising India',
            'The Great India Basket',
            'Sankranti Premium',
            'BALANCED Premium',
            'Conservative Premium'
        ],
        'Color': ['#DC2626', '#10B981', '#10B981', '#DC2626', '#FF6B35', '#FF6B35', '#10B981', '#8B5CF6'],
        'Description': [
            'Hybrid aggressive portfolio for medium to long-term goals',
            'Conservative hybrid balanced portfolio for shorter horizons',
            'Balanced diversified portfolio for systematic investment with controlled risk',
            'High-growth aggressive SIP portfolio focused on infrastructure and thematic opportunities',
            'Diversified equity basket capturing Indias growth story',
            'Diversified multi-asset allocation basket for balanced growth and stability',
            'Balanced premium portfolio for moderate risk and steady growth',
            'Conservative balanced portfolio for capital preservation with steady growth'
        ],
        'AgeRange': ['All ages', 'All ages', '25-60 years', '25-50 years', 'All ages', 'All ages', 'All ages', 'All ages'],
        'RiskLevel': ['High', 'Low', 'Medium', 'High', 'Medium-High', 'Medium', 'Medium', 'Low'],
        'MinInvestment': [15000, 10000, 5000, 10000, 10000, 10000, 10000, 10000],
        'TimeHorizon': ['3-5 years', '1-3 years', '5+ years', '7+ years', '5-7 years', '3-5 years', '3-5 years', '3-5 years'],
        'CAGR1Y': [6.62, 11.13, 16.5, 18.2, 3.4, 8.63, 6.21, 9.18],
        'CAGR3Y': [15.6, 12.32, 15.2, 16.8, 20.42, 17.69, 14.99, 17.36],
        'CAGR5Y': [16.91, 10.43, 14.8, 15.5, 25.01, 17.94, 14.69, 18.26],
        'RiskPercentage': [7.59, 5.89, 15.5, 18.5, 12.5, 14.5, 12.0, 10.5],
        'SharpeRatio': [0.31, 0.21, 0.65, 0.45, 0.85, 1.15, 1.08, 1.25],
        'ExperienceLevel': [
            'Standard to Expert',
            'Beginner to Expert',
            'Beginner to Expert',
            'Intermediate to Expert',
            'Intermediate to Expert',
            'Beginner to Expert',
            'Beginner to Expert',
            'Beginner to Expert'
        ],
        'Goals': [
            'Wealth Creation|Medium-term Goals|Marriage|Vehicle',
            'Short-term Goals|Parking Fund|Emergency Buffer|Marriage',
            'SIP Wealth Creation|Retirement|Long-term Goals|Child Education',
            'Aggressive SIP Growth|Thematic Investment|Wealth Maximization',
            'Long-term Wealth|Retirement|Children Education|Dream Home',
            'Wealth Creation|Retirement|Child Education|Balanced Growth',
            'Wealth Creation|Retirement|Balanced Growth|Capital Preservation',
            'Capital Preservation|Retirement|Stable Income|Wealth Protection'
        ],
        'ExcelFileName': [
            'Aggressive Premium.xlsx',
            'Balanced basket.xlsx',
            'every common india basket.xlsx',
            'Raising_India UPDATED.xlsx',
            'the great india basket.xlsx',
            'SankrantiPremium.xlsx',
            'BALANCED Premium.xlsx',
            'Conservative Premium.xlsx'
        ],
        'RebalancingFrequency': ['Quarterly', 'Half-yearly', 'Annually', 'Half-yearly', 'Quarterly', 'Quarterly', 'Quarterly', 'Half-yearly']
    }
    
    # Sheet 2: FundAllocations - Sample data for all baskets
    fund_allocations_data = {
        'BasketID': [
            'b9', 'b9', 'b9', 'b9', 'b9', 'b9',
            'b10', 'b10', 'b10', 'b10',
            'b12', 'b12', 'b12', 'b12',
            'b13', 'b13', 'b13', 'b13', 'b13',
            'b14', 'b14', 'b14', 'b14',
            'b16', 'b16', 'b16', 'b16', 'b16', 'b16',
            'b17', 'b17', 'b17', 'b17', 'b17', 'b17',
            'b18', 'b18', 'b18', 'b18', 'b18'
        ],
        'FundID': [
            'fund-agg-1', 'fund-agg-2', 'fund-agg-3', 'fund-agg-4', 'fund-agg-5', 'fund-agg-6',
            'fund-bal-1', 'fund-bal-2', 'fund-bal-3', 'fund-bal-4',
            'fund-eci-1', 'fund-eci-2', 'fund-eci-3', 'fund-eci-4',
            'fund-ri-1', 'fund-ri-2', 'fund-ri-3', 'fund-ri-4', 'fund-ri-5',
            'fund-gib-1', 'fund-gib-2', 'fund-gib-3', 'fund-gib-4',
            'fund-sp-1', 'fund-sp-2', 'fund-sp-3', 'fund-sp-4', 'fund-sp-5', 'fund-sp-6',
            'fund-bp-1', 'fund-bp-2', 'fund-bp-3', 'fund-bp-4', 'fund-bp-5', 'fund-bp-6',
            'fund-cp-1', 'fund-cp-2', 'fund-cp-3', 'fund-cp-4', 'fund-cp-5'
        ],
        'FundName': [
            'ICICI Prudential Active Momentum Fund',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'ICICI Prudential Technology Fund',
            'SBI Consumption Opportunities Fund',
            'Mirae Asset Nifty Smallcap 250 Momentum ETF FOF',
            'DSP Multi Asset Allocation Fund',
            'NIPPON INDIA MULTI ASSET ALLOCATION FUND',
            'DSP Small Cap Fund',
            'Kotak Large & Midcap Fund',
            'Aditya Birla Sun Life Flexi Cap Fund',
            'ICICI Pru Multi Asset Fund',
            'Quant Small Cap Fund',
            'DSP Large & Mid Cap Fund',
            'Kotak Flexicap Fund',
            'Bandhan Small Cap Fund',
            'HDFC Non-Cyclical Consumer Fund',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'Tata Digital India Fund',
            'Axis Mid Cap Fund',
            'DSP Multi Asset Allocation Fund',
            'HDFC Small Cap Fund',
            'ICICI Prudential Large & Mid Cap Fund',
            'Aditya Birla Sun Life Flexi Cap Fund',
            'ICICI Prudential Multi-Asset Fund',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'SBI Technology Opportunities Fund',
            'Aditya Birla Sun Life Consumption Fund',
            'Mirae Asset Nifty MidSmallcap400 Momentum ETF FOF',
            'ICICI Prudential Exports and Services Fund',
            'ICICI Prudential Banking and Financial Services Fund',
            'NIPPON INDIA MULTI ASSET ALLOCATION FUND',
            'ICICI Prudential FMCG Fund',
            'Mirae Asset Nifty MidSmallcap400 Momentum ETF FOF',
            'ICICI Prudential Exports and Services Fund',
            'Aditya Birla Sun Life Digital India Fund',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'SBI Technology Opportunities Fund',
            'Mirae Asset Great Consumer Fund',
            'ICICI Prudential Multi-Asset Fund',
            'ICICI Prudential Nifty Alpha Low-Volatility 30 ETF FOF'
        ],
        'AllocationPercent': [
            15, 20, 20, 15, 15, 15,
            25, 25, 25, 25,
            25, 25, 25, 25,
            15.42, 14.96, 5.88, 4.38, 3.55,
            25, 25, 25, 25,
            20, 20, 20, 15, 15, 10,
            20, 20, 15, 15, 10, 20,
            20, 20, 20, 20, 20
        ],
        'Category': [
            'Sectoral/Thematic', 'Sectoral Banking', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Fund of Funds', 'Multi Asset Allocation',
            'Multi Asset Allocation', 'Small Cap Fund', 'Large & Mid Cap', 'Flexi Cap Fund',
            'Multi Asset', 'Small Cap', 'Large & Mid Cap', 'Flexi Cap',
            'Small Cap', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Mid Cap',
            'Multi Asset Allocation', 'Small Cap Fund', 'Large & Mid Cap Fund', 'Flexi Cap Fund',
            'Multi Asset Allocation', 'Sectoral Banking', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Fund of Funds', 'Sectoral/Thematic',
            'Sectoral Banking', 'Multi Asset Allocation', 'Sectoral/Thematic', 'Fund of Funds', 'Sectoral/Thematic', 'Sectoral/Thematic',
            'Sectoral Banking', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Multi Asset Allocation', 'Fund of Funds'
        ]
    }
    
    # Sheet 3: SectorAllocations - Top sectors for each basket
    sector_allocations_data = {
        'BasketID': [
            'b9', 'b9', 'b9', 'b9', 'b9',
            'b10', 'b10', 'b10', 'b10', 'b10',
            'b12', 'b12', 'b12', 'b12', 'b12',
            'b13', 'b13', 'b13', 'b13', 'b13',
            'b14', 'b14', 'b14', 'b14', 'b14',
            'b16', 'b16', 'b16', 'b16', 'b16',
            'b17', 'b17', 'b17', 'b17', 'b17',
            'b18', 'b18', 'b18', 'b18', 'b18'
        ],
        'SectorName': [
            'Banks', 'IT Services', 'Automobiles', 'Wireless Telecom', 'Insurance',
            'Banks', 'Auto Components', 'IT Services', 'Chemicals', 'Pharmaceuticals',
            'Banks', 'Pharmaceuticals', 'IT Services', 'Automobiles', 'Real Estate',
            'IT Services', 'Banks', 'Hotels & Leisure', 'Capital Markets', 'Chemicals',
            'Banks', 'IT Services', 'Auto Components', 'Automobiles', 'Real Estate',
            'Banks', 'IT Services', 'Insurance', 'Wireless Telecom', 'Automobiles',
            'Banks', 'IT Services', 'Tobacco', 'Insurance', 'Food Products',
            'Banks', 'IT Services', 'Automobiles', 'Wireless Telecom', 'Insurance'
        ],
        'Percent': [
            14.6, 13.12, 4.77, 4.43, 4.26,
            13.43, 5.89, 5.87, 4.81, 4.3,
            18.28, 5.03, 4.26, 4.08, 3.82,
            15.42, 14.96, 5.88, 4.38, 3.55,
            13.39, 7.83, 4.92, 4.51, 4.46,
            18.43, 10.76, 4.87, 3.72, 2.99,
            16.67, 14.1, 4.83, 4.55, 3.41,
            14.82, 9.5, 4.92, 3.97, 3.43
        ]
    }
    
    # Sheet 4: TopHoldings - Top 5-10 holdings per basket
    top_holdings_data = {
        'BasketID': [
            'b9', 'b9', 'b9', 'b9', 'b9',
            'b10', 'b10', 'b10', 'b10', 'b10',
            'b12', 'b12', 'b12', 'b12', 'b12',
            'b13', 'b13', 'b13', 'b13', 'b13',
            'b14', 'b14', 'b14', 'b14', 'b14',
            'b16', 'b16', 'b16', 'b16', 'b16',
            'b17', 'b17', 'b17', 'b17', 'b17',
            'b18', 'b18', 'b18', 'b18', 'b18'
        ],
        'StockName': [
            'Infosys Limited', 'HDFC Bank Limited', 'ICICI Bank Limited', 'Bharti Airtel Limited', 'Axis Bank Limited',
            'HDFC Bank Limited', 'ICICI Bank Limited', 'Infosys Limited', 'State Bank Of India', 'Reliance Industries Limited',
            'HDFC Bank Limited', 'ICICI Bank Limited', 'Axis Bank Limited', 'Reliance Industries Limited', 'Larsen & Toubro Limited',
            'Infosys Limited', 'Eternal Limited', 'HDFC Bank Limited', 'ICICI Bank Limited', 'Tata Consultancy Services Limited',
            'ICICI Bank Limited', 'HDFC Bank Limited', 'Infosys Limited', 'Axis Bank Limited', 'Maruti Suzuki India Limited',
            'ICICI Bank Limited', 'HDFC Bank Limited', 'Infosys Limited', 'Bharti Airtel Limited', 'Axis Bank Limited',
            'HDFC Bank Limited', 'Infosys Limited', 'ICICI Bank Limited', 'ITC Limited', 'Hindustan Unilever Limited',
            'Bharti Airtel Limited', 'HDFC Bank Limited', 'ICICI Bank Limited', 'Infosys Limited', 'Axis Bank Limited'
        ],
        'Percent': [
            4.63, 4.53, 3.69, 2.89, 2.16,
            3.57, 3.19, 2.3, 2.22, 1.95,
            4.36, 4.15, 2.82, 1.85, 1.74,
            3.78, 3.33, 3.12, 2.96, 2.24,
            3.05, 2.65, 2.31, 2.23, 1.57,
            4.92, 4.89, 4.42, 3.72, 2.85,
            5.4, 5.02, 4.94, 4.83, 2.9,
            3.97, 3.74, 3.67, 3.48, 2.43
        ],
        'Sector': [
            'IT Services', 'Banks', 'Banks', 'Wireless Telecom', 'Banks',
            'Banks', 'Banks', 'IT Services', 'Banks', 'Oil & Gas',
            'Banks', 'Banks', 'Banks', 'Oil & Gas', 'Construction',
            'IT Services', 'Chemicals', 'Banks', 'Banks', 'IT Services',
            'Banks', 'Banks', 'IT Services', 'Banks', 'Automobiles',
            'Banks', 'Banks', 'IT Services', 'Wireless Telecom', 'Banks',
            'Banks', 'IT Services', 'Banks', 'Tobacco', 'Household Products',
            'Wireless Telecom', 'Banks', 'Banks', 'IT Services', 'Banks'
        ],
        'Type': [
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity',
            'Equity', 'Equity', 'Equity', 'Equity', 'Equity'
        ]
    }
    
    # Sheet 5: Funds - Master fund list
    funds_data = {
        'FundID': [
            'fund-agg-1', 'fund-agg-2', 'fund-agg-3', 'fund-agg-4', 'fund-agg-5', 'fund-agg-6',
            'fund-bal-1', 'fund-bal-2', 'fund-bal-3', 'fund-bal-4',
            'fund-eci-1', 'fund-eci-2', 'fund-eci-3', 'fund-eci-4',
            'fund-ri-1', 'fund-ri-2', 'fund-ri-3', 'fund-ri-4', 'fund-ri-5'
        ],
        'Name': [
            'ICICI Prudential Active Momentum Fund',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'ICICI Prudential Technology Fund',
            'SBI Consumption Opportunities Fund',
            'Mirae Asset Nifty Smallcap 250 Momentum ETF FOF',
            'DSP Multi Asset Allocation Fund',
            'NIPPON INDIA MULTI ASSET ALLOCATION FUND',
            'DSP Small Cap Fund',
            'Kotak Large & Midcap Fund',
            'Aditya Birla Sun Life Flexi Cap Fund',
            'ICICI Pru Multi Asset Fund',
            'Quant Small Cap Fund',
            'DSP Large & Mid Cap Fund',
            'Kotak Flexicap Fund',
            'Bandhan Small Cap Fund',
            'HDFC Non-Cyclical Consumer Fund',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'Tata Digital India Fund',
            'Axis Mid Cap Fund'
        ],
        'Category': [
            'Sectoral/Thematic', 'Sectoral Banking', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Fund of Funds', 'Multi Asset Allocation',
            'Multi Asset Allocation', 'Small Cap Fund', 'Large & Mid Cap', 'Flexi Cap Fund',
            'Multi Asset', 'Small Cap', 'Large & Mid Cap', 'Flexi Cap',
            'Small Cap', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Sectoral/Thematic', 'Mid Cap'
        ],
        'NAV': [
            10.54, 667.25, 211.14, 296.21, 9.46, 15.92,
            24.24, 125.45, 89.32, 678.91,
            145.23, 78.56, 234.67, 156.78,
            198.45, 145.67, 667.25, 89.34, 112.56
        ],
        'Returns1Y': [
            0.0, 20.79, -0.67, -4.95, -8.75, 26.68,
            22.38, 15.23, 18.45, 21.34,
            16.89, 19.45, 17.23, 18.67,
            24.56, 18.92, 20.79, 16.45, 22.34
        ],
        'Returns3Y': [
            0.0, 17.94, 15.69, 13.16, -2.89, 21.76,
            20.84, 18.67, 16.45, 19.23,
            19.36, 21.45, 18.23, 19.67,
            22.45, 20.15, 17.94, 18.23, 21.56
        ],
        'Returns5Y': [
            0.0, 19.41, 15.21, 16.1, 0.0, 0.0,
            16.61, 17.89, 15.67, 18.45,
            20.63, 19.78, 17.45, 18.89,
            21.34, 19.45, 19.41, 17.89, 20.23
        ],
        'ExpenseRatio': [
            1.2, 1.15, 1.18, 1.22, 1.45, 1.35,
            1.25, 1.38, 1.28, 1.32,
            1.15, 1.42, 1.28, 1.25,
            1.45, 1.32, 1.15, 1.35, 1.38
        ],
        'AUM': [
            50, 7791, 15565, 3218, 212, 5586,
            9600, 4523, 8901, 12456,
            75067, 3456, 6789, 9876,
            5678, 3456, 7791, 4567, 6789
        ],
        'Risk': [
            'High', 'High', 'High', 'High', 'High', 'Medium',
            'Medium', 'High', 'Medium-High', 'Medium-High',
            'Medium', 'High', 'Medium-High', 'Medium-High',
            'High', 'High', 'High', 'High', 'High'
        ],
        'Rating': [
            3, 5, 4, 4, 3, 5,
            5, 4, 5, 4,
            5, 4, 5, 4,
            4, 4, 5, 4, 5
        ]
    }
    
    # Create DataFrames
    df_baskets = pd.DataFrame(baskets_data)
    df_fund_allocations = pd.DataFrame(fund_allocations_data)
    df_sector_allocations = pd.DataFrame(sector_allocations_data)
    df_top_holdings = pd.DataFrame(top_holdings_data)
    df_funds = pd.DataFrame(funds_data)
    
    # Write to Excel
    print(f"📝 Creating Excel file: {excel_path}")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_baskets.to_excel(writer, sheet_name='Baskets', index=False)
        df_fund_allocations.to_excel(writer, sheet_name='FundAllocations', index=False)
        df_sector_allocations.to_excel(writer, sheet_name='SectorAllocations', index=False)
        df_top_holdings.to_excel(writer, sheet_name='TopHoldings', index=False)
        df_funds.to_excel(writer, sheet_name='Funds', index=False)
    
    # Format Excel
    print("🎨 Formatting Excel sheets...")
    wb = load_workbook(excel_path)
    
    # Style headers for all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    
    print(f"\n✅ Excel template created successfully!")
    print(f"📊 Location: {excel_path}")
    print(f"\n📈 Summary:")
    print(f"   • Baskets: {len(df_baskets)} (b9, b10, b12, b13, b14, b16, b17, b18)")
    print(f"   • Fund Allocations: {len(df_fund_allocations)} entries")
    print(f"   • Sector Allocations: {len(df_sector_allocations)} entries")
    print(f"   • Top Holdings: {len(df_top_holdings)} entries")
    print(f"   • Funds: {len(df_funds)} unique funds")
    print(f"\n💡 Note: You can easily add more baskets by adding rows to the Baskets sheet!")
    print(f"   Just copy the format and fill in the data - changes will reflect in 2 minutes!")

if __name__ == "__main__":
    create_complete_excel()
