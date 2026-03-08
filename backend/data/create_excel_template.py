"""
Script to create the master Excel template for Alphanifty basket data
Run this once to generate AlphaniftyMasterData.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_master_excel():
    """Create the master Excel file with all sheets and sample data"""
    
    excel_path = os.path.join(os.path.dirname(__file__), 'AlphaniftyMasterData.xlsx')
    
    # Sheet 1: Baskets
    baskets_data = {
        'BasketID': ['b4', 'b9', 'b10'],
        'Name': ['Yellow Basket', 'Aggressive Premium', 'Conservative Balanced Basket'],
        'Color': ['#E8C23A', '#DC2626', '#10B981'],
        'Description': [
            'Conservative equity savings for very short-term goals',
            'Hybrid aggressive portfolio for medium to long-term goals',
            'Conservative hybrid balanced portfolio for shorter horizons'
        ],
        'AgeRange': ['All ages', 'All ages', 'All ages'],
        'RiskLevel': ['Low', 'High', 'Low'],
        'MinInvestment': [10000, 15000, 10000],
        'TimeHorizon': ['Less than 1 year', '3-5 years', '1-3 years'],
        'CAGR1Y': [11.8, 6.62, 11.13],
        'CAGR3Y': [10.3, 15.6, 12.32],
        'CAGR5Y': [9.2, 16.91, 10.43],
        'RiskPercentage': [8.6, 7.59, 5.89],
        'SharpeRatio': [1.37, 0.31, 0.21],
        'ExperienceLevel': ['Beginner to Expert', 'Standard to Expert', 'Beginner to Expert'],
        'Goals': [
            'Short-term Goals|Emergency Fund|Parking Fund',
            'Wealth Creation|Medium-term Goals|Marriage',
            'Short-term Goals|Parking Fund|Emergency Buffer'
        ],
        'ExcelFileName': ['Yellow basket.xlsx', 'Aggressive Premium.xlsx', 'Balanced basket.xlsx'],
        'Philosophy': [
            'Short-term investing requires maximum safety with modest growth potential',
            'Aggressive hybrid investing bridges the gap between pure equity and balanced approaches',
            'Conservative investing does not mean sacrificing returns'
        ],
        'RebalancingFrequency': ['Quarterly', 'Quarterly', 'Half-yearly']
    }
    
    # Sheet 2: FundAllocations
    fund_allocations_data = {
        'BasketID': ['b4', 'b4', 'b4', 'b9', 'b9', 'b9'],
        'FundID': ['fund-16', 'fund-17', 'fund-18', 'fund-agg-1', 'fund-agg-2', 'fund-agg-3'],
        'FundName': [
            'Kotak Equity Savings Fund(G)',
            'HDFC Equity Savings Fund(G)',
            'ICICI Pru Equity Savings Fund-Reg(G)',
            'ICICI Prudential Active Momentum Fund - Growth',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND',
            'ICICI Prudential Technology Fund - Growth'
        ],
        'AllocationPercent': [35, 35, 30, 15, 20, 20],
        'Category': [
            'Equity Savings',
            'Equity Savings',
            'Equity Savings',
            'Sectoral/Thematic',
            'Sectoral Banking and Financial',
            'Sectoral/Thematic'
        ]
    }
    
    # Sheet 3: SectorAllocations
    sector_allocations_data = {
        'BasketID': ['b9', 'b9', 'b9', 'b9', 'b9', 'b10', 'b10', 'b10'],
        'SectorName': ['Banks', 'IT Services', 'Automobiles', 'Wireless Telecom', 'Insurance', 'Banks', 'Auto Components', 'IT Services'],
        'Percent': [14.6, 13.12, 4.77, 4.43, 4.26, 13.43, 5.89, 5.87]
    }
    
    # Sheet 4: TopHoldings
    top_holdings_data = {
        'BasketID': ['b9', 'b9', 'b9', 'b9', 'b10', 'b10', 'b10'],
        'StockName': [
            'Infosys Limited',
            'HDFC Bank Limited',
            'ICICI Bank Limited',
            'Bharti Airtel Limited',
            'HDFC Bank Limited',
            'ICICI Bank Limited',
            'Infosys Limited'
        ],
        'Percent': [4.63, 4.53, 3.69, 2.89, 3.57, 3.19, 2.3],
        'Sector': ['IT Services', 'Banks', 'Banks', 'Wireless Telecom', 'Banks', 'Banks', 'IT Services'],
        'Type': ['Equity', 'Equity', 'Equity', 'Equity', 'Equity', 'Equity', 'Equity']
    }
    
    # Sheet 5: Funds
    funds_data = {
        'FundID': ['fund-16', 'fund-17', 'fund-18', 'fund-agg-1', 'fund-agg-2'],
        'Name': [
            'Kotak Equity Savings Fund(G)',
            'HDFC Equity Savings Fund(G)',
            'ICICI Pru Equity Savings Fund-Reg(G)',
            'ICICI Prudential Active Momentum Fund - Growth',
            'NIPPON INDIA BANKING & FINANCIAL SERVICES FUND'
        ],
        'Category': ['Equity Savings', 'Equity Savings', 'Equity Savings', 'Sectoral/Thematic', 'Sectoral Banking'],
        'NAV': [18.45, 55.32, 198.67, 10.54, 667.25],
        'Returns1Y': [9.8, 10.2, 8.9, 0.0, 20.79],
        'Returns3Y': [8.5, 9.1, 8.2, 0.0, 17.94],
        'Returns5Y': [7.8, 8.4, 7.5, 0.0, 19.41],
        'ExpenseRatio': [0.85, 0.78, 0.92, 1.2, 1.15],
        'AUM': [5200, 8900, 12500, 50, 7791],
        'Risk': ['Low', 'Low', 'Low', 'High', 'High'],
        'Rating': [4, 5, 4, 3, 5]
    }
    
    # Create DataFrames
    df_baskets = pd.DataFrame(baskets_data)
    df_fund_allocations = pd.DataFrame(fund_allocations_data)
    df_sector_allocations = pd.DataFrame(sector_allocations_data)
    df_top_holdings = pd.DataFrame(top_holdings_data)
    df_funds = pd.DataFrame(funds_data)
    
    # Write to Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_baskets.to_excel(writer, sheet_name='Baskets', index=False)
        df_fund_allocations.to_excel(writer, sheet_name='FundAllocations', index=False)
        df_sector_allocations.to_excel(writer, sheet_name='SectorAllocations', index=False)
        df_top_holdings.to_excel(writer, sheet_name='TopHoldings', index=False)
        df_funds.to_excel(writer, sheet_name='Funds', index=False)
    
    # Format Excel
    wb = load_workbook(excel_path)
    
    # Style headers for all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    
    print(f"✅ Excel template created successfully at: {excel_path}")
    print(f"\n📊 Sheets created:")
    print(f"   1. Baskets - {len(df_baskets)} baskets")
    print(f"   2. FundAllocations - {len(df_fund_allocations)} allocations")
    print(f"   3. SectorAllocations - {len(df_sector_allocations)} sectors")
    print(f"   4. TopHoldings - {len(df_top_holdings)} holdings")
    print(f"   5. Funds - {len(df_funds)} funds")
    print(f"\n💡 You can now edit this file to manage basket data!")
    
    return excel_path

if __name__ == "__main__":
    create_master_excel()
